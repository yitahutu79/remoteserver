# 按照关键词对表单进行分类，归类到不同商品表中

import openpyxl

def main(path):
    # 得到sheet页
    wb = openpyxl.load_workbook(path)
    table = wb.worksheets[0]

    wb_list=[]
    sheet_title=['小麦胚芽','布丁粉','双皮奶','烧仙草','椰奶冻','五彩面粉','燕麦','黑芝麻糊','藕粉速溶','杏仁粉','阿萨姆奶茶','豆浆','缤纷谷物','奶盖粉','黄瓜籽粉','批发专用','拉丝酸奶粉','粽叶','肠衣','五香卤蛋调料包','调味料',
    '红枣粉','山药百合牛乳','山药粉','九红粉','其他']

    # 删除多余sheet
    sheets = wb.sheetnames[1:]
    for sheet in sheets:
        del wb[sheet]

    # for i in range(len(sheet_title)):
    #     if sheet_title[i] in wb.sheetnames:
    #         print("该表已存在")
    #     else:
    #         print("正在创建表"+sheet_title[i])
    #         wb.create_sheet(sheet_title[i])
    #         wb_list.append(wb[sheet_title[i]])

    # 总行数
    nrows=table.max_row
    # 关键词
    keyword=['小麦胚芽','布丁','双皮奶','仙草','椰奶冻','面粉','麦','芝麻糊','藕粉','杏仁','奶茶','豆浆','谷物','奶盖','黄瓜','批发','酸奶','粽叶','肠','调料包','调味料','红枣','牛乳','山药','九红']


    # 按照关键词分类存放到不同sheet表
    for i in range(1,nrows+1):
        flag = 0
        for j in range(len(keyword)):
            # # 列名整体拷贝
            # if i == 1:
            #     wb_list[j].append([cell.value for cell in table[i]])
            # elif (flag == 0) & ((keyword[j] in table.cell(i,3).value) | (keyword[j] in table.cell(i,6).value)):
            #     flag = 1
            #     wb_list[j].append([cell.value for cell in table[i]])
            if (flag == 0) & ((keyword[j] in table.cell(i,1).value) | (keyword[j] in table.cell(i,5).value)):
                flag = 1
                if not sheet_title[j] in wb.sheetnames:
                    wb.create_sheet(sheet_title[j])
                    wb[sheet_title[j]].append([cell.value for cell in table[1]])
                wb[sheet_title[j]].append([cell.value for cell in table[i]])

        if flag == 0:
            # wb_list[-1].append([cell.value for cell in table[i]])
            if not sheet_title[-1] in wb.sheetnames:
                wb.create_sheet(sheet_title[-1])
            wb[sheet_title[-1]].append([cell.value for cell in table[i]])

    wb.save(path)

if __name__ == '__main__':
    # path='/root/autodl-tmp/lj_workplace/0110/iter/data/副本沁润 qa整理.xlsx'
    # main(path)
    path='/root/autodl-tmp/lj_workplace/0110/iter/data/副本沁润_快捷短语.xlsx'
    main(path)